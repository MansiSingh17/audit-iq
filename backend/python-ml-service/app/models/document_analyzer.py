from transformers import T5Tokenizer, T5ForConditionalGeneration
import torch
import PyPDF2
import docx
from openpyxl import load_workbook
import base64
import io
from app.config import Config
from app.utils.logger import get_logger

logger = get_logger(__name__)

class DocumentAnalyzer:
    """Document text extraction and summarization"""
    
    def __init__(self):
        self.model_name = Config.MODEL_NAME
        self.device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
        logger.info(f'Loading T5 model for summarization: {self.model_name}')
        
        self.tokenizer = T5Tokenizer.from_pretrained(self.model_name)
        self.model = T5ForConditionalGeneration.from_pretrained(self.model_name)
        self.model.to(self.device)
        self.model.eval()
        
        logger.info('Document analyzer initialized successfully')
    
    def extract_text(self, file_content, file_type):
        """Extract text from various document formats"""
        try:
            logger.info(f'Extracting text from {file_type}')
            
            # Decode base64 content
            file_bytes = base64.b64decode(file_content)
            file_stream = io.BytesIO(file_bytes)
            
            if file_type == 'pdf':
                text = self._extract_from_pdf(file_stream)
            elif file_type in ['docx', 'doc']:
                text = self._extract_from_docx(file_stream)
            elif file_type in ['xlsx', 'xls']:
                text = self._extract_from_excel(file_stream)
            else:
                text = file_bytes.decode('utf-8', errors='ignore')
            
            result = {
                'success': True,
                'text': text,
                'length': len(text),
                'file_type': file_type
            }
            
            logger.info(f'Extracted {len(text)} characters')
            return result
            
        except Exception as e:
            logger.error(f'Error extracting text: {str(e)}')
            return {
                'success': False,
                'error': str(e),
                'text': ''
            }
    
    def _extract_from_pdf(self, file_stream):
        """Extract text from PDF"""
        reader = PyPDF2.PdfReader(file_stream)
        text = ''
        for page in reader.pages:
            text += page.extract_text() + '\n'
        return text.strip()
    
    def _extract_from_docx(self, file_stream):
        """Extract text from DOCX"""
        doc = docx.Document(file_stream)
        text = '\n'.join([para.text for para in doc.paragraphs])
        return text.strip()
    
    def _extract_from_excel(self, file_stream):
        """Extract text from Excel"""
        workbook = load_workbook(file_stream)
        text = ''
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                text += ' '.join([str(cell) for cell in row if cell]) + '\n'
        return text.strip()
    
    def summarize(self, text):
        """Summarize document text using T5"""
        try:
            logger.info('Summarizing text')
            
            # Prepare input
            input_text = f"summarize: {text[:2000]}"  # Limit to 2000 chars
            inputs = self.tokenizer(
                input_text,
                max_length=512,
                truncation=True,
                return_tensors='pt'
            ).to(self.device)
            
            # Generate summary
            with torch.no_grad():
                summary_ids = self.model.generate(
                    inputs['input_ids'],
                    max_length=150,
                    min_length=50,
                    length_penalty=2.0,
                    num_beams=4,
                    early_stopping=True
                )
            
            summary = self.tokenizer.decode(summary_ids[0], skip_special_tokens=True)
            
            result = {
                'success': True,
                'original_length': len(text),
                'summary': summary,
                'summary_length': len(summary),
                'compression_ratio': round(len(summary) / len(text), 2)
            }
            
            logger.info(f'Generated summary: {len(summary)} characters')
            return result
            
        except Exception as e:
            logger.error(f'Error summarizing text: {str(e)}')
            raise